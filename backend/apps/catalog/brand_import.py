"""品牌 + 系列 批次匯入 service。

CSV / Excel 格式(一行一個系列,同品牌有多列):

    品牌名稱 | 品牌代碼(選填) | 系列名稱(選填) | 系列代碼(選填) | 品牌排序(選填) | 系列排序(選填)

範例:
    Apple,apple,iPhone,iphone,1,1
    Apple,apple,iPad,ipad,1,2
    Samsung,samsung,Galaxy S,s,2,1

- 品牌名稱必填;代碼留空 → 自動 slugify
- 系列名稱留空 → 該列只新增/更新品牌
- 已存在的 (品牌 code) 會 update_or_create
- 已存在的 (品牌, 系列 code) 會 update_or_create
"""
import csv
import io
import re
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from .models import Brand, PhoneSeries


_FIELD_MAP: dict[str, str] = {
    "品牌": "brand_name",
    "品牌名稱": "brand_name",
    "brand": "brand_name",
    "brand_name": "brand_name",
    "品牌代碼": "brand_code",
    "brand_code": "brand_code",
    "系列": "series_name",
    "系列名稱": "series_name",
    "series": "series_name",
    "series_name": "series_name",
    "系列代碼": "series_code",
    "series_code": "series_code",
    "品牌排序": "brand_sort",
    "brand_sort": "brand_sort",
    "系列排序": "series_sort",
    "series_sort": "series_sort",
}


def _norm(h: Any) -> str:
    return str(h or "").strip().lower()


def _slugify(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^a-z0-9\-]", "", s)
    return s[:20]


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    header = rows[0]
    return [dict(zip(header, r)) for r in rows[1:] if any(r)]


def _parse_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "") for h in rows[0]]
    return [
        dict(zip(header, [str(c) if c is not None else "" for c in r]))
        for r in rows[1:]
        if any(c is not None and str(c).strip() for c in r)
    ]


def _normalize(raw: dict) -> dict:
    out: dict = {}
    for k, v in raw.items():
        key = _FIELD_MAP.get(_norm(k))
        if key:
            out[key] = str(v or "").strip()
    return out


def parse_file(file_obj, filename: str) -> list[dict]:
    """根據副檔名解析 → 統一格式 dict 清單。"""
    content = file_obj.read()
    if filename.lower().endswith(".xlsx"):
        raw_rows = _parse_xlsx(content)
    else:
        raw_rows = _parse_csv(content)
    return [_normalize(r) for r in raw_rows if r]


def import_brands_series(
    tenant, file_obj, filename: str, dry_run: bool = True
) -> dict:
    """匯入品牌與系列。

    回傳:
        {
          "summary": { brands_created, brands_updated, series_created,
                       series_updated, rows_skipped },
          "preview": [...],
          "errors": [...]
        }
    """
    rows = parse_file(file_obj, filename)
    if not rows:
        return {
            "summary": {
                "brands_created": 0, "brands_updated": 0,
                "series_created": 0, "series_updated": 0,
                "rows_skipped": 0,
            },
            "preview": [],
            "errors": [{"line": 0, "msg": "檔案沒有可讀取的資料列"}],
        }

    existing_brands = {b.code: b for b in Brand.objects.for_tenant(tenant)}
    existing_series_by_brand: dict[str, dict[str, PhoneSeries]] = {}

    preview: list[dict] = []
    errors: list[dict] = []
    brands_new = 0
    brands_upd = 0
    series_new = 0
    series_upd = 0
    skipped = 0

    # 暫存要寫入的資料
    to_write_brands: dict[str, dict] = {}  # code → {name, sort}
    to_write_series: list[dict] = []  # [{brand_code, code, name, sort}]

    for idx, r in enumerate(rows, start=2):  # 第 2 行起 (header 是 1)
        brand_name = r.get("brand_name", "").strip()
        if not brand_name:
            skipped += 1
            errors.append({"line": idx, "msg": "品牌名稱為空"})
            continue
        brand_code = r.get("brand_code", "").strip() or _slugify(brand_name)
        series_name = r.get("series_name", "").strip()
        series_code = r.get("series_code", "").strip()
        if series_name and not series_code:
            series_code = _slugify(series_name)

        brand_sort_str = r.get("brand_sort", "").strip()
        series_sort_str = r.get("series_sort", "").strip()
        try:
            brand_sort = int(brand_sort_str) if brand_sort_str else 99
        except ValueError:
            brand_sort = 99
        try:
            series_sort = int(series_sort_str) if series_sort_str else 99
        except ValueError:
            series_sort = 99

        existing = existing_brands.get(brand_code)
        brand_action = "(已存在)" if existing else "新增"
        if not existing:
            brands_new += 1
        else:
            brands_upd += 1
        to_write_brands[brand_code] = {
            "name": brand_name,
            "sort": brand_sort,
        }

        if series_name:
            # series action
            if brand_code not in existing_series_by_brand:
                if existing:
                    existing_series_by_brand[brand_code] = {
                        s.code: s for s in existing.series.all()
                    }
                else:
                    existing_series_by_brand[brand_code] = {}
            existing_series = existing_series_by_brand[brand_code].get(
                series_code
            )
            series_action = "(已存在)" if existing_series else "新增"
            if existing_series:
                series_upd += 1
            else:
                series_new += 1
            to_write_series.append(
                {
                    "brand_code": brand_code,
                    "code": series_code,
                    "name": series_name,
                    "sort": series_sort,
                }
            )
        else:
            series_action = "—(只匯入品牌)"

        preview.append(
            {
                "line": idx,
                "brand_name": brand_name,
                "brand_code": brand_code,
                "brand_action": brand_action,
                "series_name": series_name,
                "series_code": series_code,
                "series_action": series_action,
            }
        )

    summary = {
        "brands_created": brands_new,
        "brands_updated": brands_upd,
        "series_created": series_new,
        "series_updated": series_upd,
        "rows_skipped": skipped,
    }

    if dry_run:
        return {"summary": summary, "preview": preview, "errors": errors}

    # ── commit
    if errors:
        return {
            "summary": summary,
            "preview": preview,
            "errors": errors,
            "detail": "驗證有錯誤,請修正後再匯入",
        }
    with transaction.atomic():
        # Brand upsert
        brand_obj_by_code: dict[str, Brand] = {}
        for code, data in to_write_brands.items():
            obj, _created = Brand.objects.update_or_create(
                tenant=tenant,
                code=code,
                defaults={"name": data["name"], "sort_order": data["sort"]},
            )
            brand_obj_by_code[code] = obj
        # Series upsert
        for s in to_write_series:
            brand = brand_obj_by_code.get(s["brand_code"])
            if brand is None:
                continue
            PhoneSeries.objects.update_or_create(
                tenant=tenant,
                brand=brand,
                code=s["code"],
                defaults={"name": s["name"], "sort_order": s["sort"]},
            )
    return {"summary": summary, "preview": preview, "errors": errors}
