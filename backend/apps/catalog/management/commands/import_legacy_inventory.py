"""舊系統「庫存明細表」匯入指令。

用法(先預覽,確認數字後再正式寫入):

    # 1) 預覽(不會改任何資料)
    python manage.py import_legacy_inventory \
        --xls "/path/庫存明細表.xls" \
        --mapping "/path/庫存匯入_類別對照表.xlsx"

    # 2) 正式執行(會先清空再匯入,全程包在一個交易裡)
    python manage.py import_legacy_inventory \
        --xls "/path/庫存明細表.xls" \
        --mapping "/path/庫存匯入_類別對照表.xlsx" --confirm

規則(已與使用者確認):
- 清空:現有銷貨/進貨/調撥單及明細、序號、庫存變動、庫存餘額、商品、類別、倉別
  全部刪除(供應商 / 客戶 / 付款方式 / 發票設定保留)。
- 排除整個「亞太巨城」倉。
- 排除對照表中標記「排除」的品名前綴(帳務 / 門號等非庫存項目)。
- 新機(品名以「手機/」開頭、無逐支 IMEI)這批不匯。
- 中古機(品名以「中古/」開頭):品號當 IMEI 建序號,一支一筆。
- 配件:依品號去重,各倉鋪在庫數量(StockBalance)。
- 類別:依對照表「你的修改」欄(空則用「提議類別」)建立。
"""
from __future__ import annotations

from collections import defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.catalog.models import Category, Product
from apps.inventory.models import (
    ProductSerial,
    StockBalance,
    StockMovement,
    Warehouse,
)
from apps.purchasing.models import PurchaseOrder, PurchaseOrderCategory, PurchaseOrderItem
from apps.sales.models import (
    SalesOrder,
    SalesOrderItem,
    SalesOrderItemSerial,
    SalesOrderPayment,
)
from apps.tenants.models import Tenant
from apps.transfers.models import (
    TransferOrder,
    TransferOrderItem,
    TransferOrderItemSerial,
)

EXCLUDE_WAREHOUSE = "亞太巨城"
NEW_PHONE_PREFIX = "手機"  # 新機這批不匯
SECONDHAND_PREFIX = "中古"

# 類別名稱 → SKU 前綴代碼(SlugField 只能 ASCII)。未列到的用 G01.. 流水。
CODE_MAP = {
    "手機": "PH",
    "中古機": "AA",
    "保護貼": "SP",
    "保護殼/背蓋": "CS",
    "皮套": "LC",
    "充電/線材": "CH",
    "耳機/音訊": "EP",
    "支架/車架": "ST",
    "掛繩/吊飾": "LN",
    "穿戴/周邊": "WB",
    "包膜": "FM",
    "維修料": "RP",
    "其他配件": "AC",
}


def parse_legacy_xls(path: str):
    """回傳 list of dict:{分類, 品號, 品名, 店名, 數量, 前綴}。"""
    from bs4 import BeautifulSoup

    raw = open(path, "rb").read().decode("big5", "replace")
    soup = BeautifulSoup(raw, "lxml")
    rows = []
    for tr in soup.find_all("tr"):
        cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
        if len(cells) >= 9 and cells[0].isdigit():
            name = cells[3]
            rows.append(
                {
                    "cat_code": cells[1],
                    "code": cells[2],
                    "name": name,
                    "store": cells[4],
                    "qty": int(cells[5]) if cells[5].isdigit() else 0,
                    "prefix": name.split("/")[0] if "/" in name else name,
                }
            )
    return rows


def load_mapping(path: str):
    """讀對照表 xlsx:回傳 {前綴: 類別名稱};標記排除者不放進來。"""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb["類別對照"] if "類別對照" in wb.sheetnames else wb.active
    header = [c.value for c in ws[1]]
    # 欄位:提議類別 / 商品型態 / 舊系統品名前綴 / 庫存筆數 / 你的修改
    idx = {name: i for i, name in enumerate(header)}
    col_prop = idx.get("提議類別", 0)
    col_prefix = idx.get("舊系統品名前綴", 2)
    col_edit = idx.get("你的修改(類別/排除)", idx.get("你的修改", 4))
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[col_prefix] in (None, ""):
            continue
        prefix = str(row[col_prefix]).strip()
        edit = (str(row[col_edit]).strip() if col_edit < len(row) and row[col_edit] not in (None, "") else "")
        proposed = str(row[col_prop]).strip() if row[col_prop] else ""
        final = edit or proposed
        if final in ("排除", "(排除)", "刪除", "不匯"):
            continue
        mapping[prefix] = final
    return mapping


class Command(BaseCommand):
    help = "匯入舊系統庫存明細表(預設預覽,--confirm 才寫入)"

    def add_arguments(self, parser):
        parser.add_argument("--xls", required=True, help="庫存明細表.xls 路徑")
        parser.add_argument("--mapping", required=True, help="類別對照表.xlsx 路徑")
        parser.add_argument("--confirm", action="store_true", help="實際清空並寫入")
        parser.add_argument("--tenant", type=int, default=1, help="租戶 ID(預設 1)")

    def handle(self, *args, **opts):
        try:
            tenant = Tenant.objects.get(id=opts["tenant"])
        except Tenant.DoesNotExist:
            raise CommandError(f"找不到 tenant id={opts['tenant']}")

        rows = parse_legacy_xls(opts["xls"])
        mapping = load_mapping(opts["mapping"])

        # 篩選要匯入的列
        kept = []
        skipped = defaultdict(int)
        for r in rows:
            if r["store"] == EXCLUDE_WAREHOUSE:
                skipped["亞太巨城"] += 1
                continue
            if r["prefix"] == NEW_PHONE_PREFIX:
                skipped["新機(不匯)"] += 1
                continue
            if r["prefix"] not in mapping:
                skipped["排除/未對應"] += 1
                continue
            if r["qty"] <= 0 and r["prefix"] != SECONDHAND_PREFIX:
                skipped["數量0"] += 1
                continue
            kept.append(r)

        warehouses = sorted({r["store"] for r in kept})
        sh_rows = [r for r in kept if r["prefix"] == SECONDHAND_PREFIX]
        acc_rows = [r for r in kept if r["prefix"] != SECONDHAND_PREFIX]
        acc_products = {r["code"] for r in acc_rows}
        categories = sorted({mapping[r["prefix"]] for r in kept})

        self.stdout.write(self.style.MIGRATE_HEADING("=== 匯入預覽 ==="))
        self.stdout.write(f"來源列(全部): {len(rows)}")
        for k, v in skipped.items():
            self.stdout.write(f"  跳過 {k}: {v}")
        self.stdout.write(f"要匯入的列: {len(kept)}")
        self.stdout.write(f"建立倉別({len(warehouses)}): {', '.join(warehouses)}")
        self.stdout.write(f"建立類別({len(categories)}): {', '.join(categories)}")
        self.stdout.write(f"中古機(序號): {len(sh_rows)} 支")
        self.stdout.write(
            f"配件商品: {len(acc_products)} 項;庫存餘額列: {len(acc_rows)}"
        )

        # 清空現況統計
        self.stdout.write(self.style.MIGRATE_HEADING("=== 將清空(目前筆數)==="))
        self.stdout.write(
            f"  銷貨單 {SalesOrder.objects.filter(tenant=tenant).count()} / "
            f"進貨單 {PurchaseOrder.objects.filter(tenant=tenant).count()} / "
            f"調撥單 {TransferOrder.objects.filter(tenant=tenant).count()}"
        )
        self.stdout.write(
            f"  序號 {ProductSerial.objects.filter(tenant=tenant).count()} / "
            f"庫存餘額 {StockBalance.objects.filter(tenant=tenant).count()} / "
            f"商品 {Product.objects.filter(tenant=tenant).count()} / "
            f"類別 {Category.objects.filter(tenant=tenant).count()} / "
            f"倉別 {Warehouse.objects.filter(tenant=tenant).count()}"
        )

        if not opts["confirm"]:
            self.stdout.write(
                self.style.WARNING(
                    "\n這是預覽,未改任何資料。確認數字無誤後,加上 --confirm 正式執行。"
                )
            )
            return

        with transaction.atomic():
            self._wipe(tenant)
            wh_map = self._create_warehouses(tenant, warehouses)
            cat_map = self._create_categories(tenant, categories)
            n_prod, n_serial, n_bal = self._import(
                tenant, kept, mapping, wh_map, cat_map
            )
        self.stdout.write(
            self.style.SUCCESS(
                f"\n完成:商品 {n_prod}、序號 {n_serial}、庫存餘額 {n_bal}。"
            )
        )

    # ---- 清空 ----
    def _wipe(self, tenant):
        f = dict(tenant=tenant)
        # 順序重點:ProductSerial 透過 PROTECT 外鍵指向 PurchaseOrderItem /
        # SalesOrder / Product / Warehouse,所以必須「先刪序號」才能刪那些;
        # 而序號又被 SalesOrderItemSerial / SalesOrderItem / StockMovement /
        # TransferOrderItemSerial PROTECT 參照,所以這些得在序號之前刪。
        SalesOrderItemSerial.objects.filter(**f).delete()
        SalesOrderPayment.objects.filter(**f).delete()
        SalesOrderItem.objects.filter(**f).delete()
        StockMovement.objects.filter(**f).delete()
        TransferOrderItemSerial.objects.filter(**f).delete()
        TransferOrderItem.objects.filter(**f).delete()
        TransferOrder.objects.filter(**f).delete()
        ProductSerial.objects.filter(**f).delete()
        PurchaseOrderItem.objects.filter(**f).delete()
        PurchaseOrder.objects.filter(**f).delete()
        SalesOrder.objects.filter(**f).delete()
        StockBalance.objects.filter(**f).delete()
        PurchaseOrderCategory.objects.filter(**f).delete()
        Product.objects.filter(**f).delete()
        Category.objects.filter(**f).delete()
        Warehouse.objects.filter(**f).delete()
        self.stdout.write("  已清空現有資料")

    # ---- 建倉 ----
    def _create_warehouses(self, tenant, names):
        wh_map = {}
        for i, name in enumerate(names, start=1):
            wh = Warehouse.objects.create(
                tenant=tenant, code=f"W{i:02d}", name=name, is_active=True
            )
            wh_map[name] = wh
        self.stdout.write(f"  建立 {len(wh_map)} 個倉別")
        return wh_map

    # ---- 建類別 ----
    def _create_categories(self, tenant, names):
        cat_map = {}
        used_codes = set()
        seq = 0
        for name in names:
            code = CODE_MAP.get(name)
            if not code or code in used_codes:
                seq += 1
                code = f"G{seq:02d}"
            used_codes.add(code)
            cat = Category.objects.create(
                tenant=tenant,
                code=code,
                name=name,
                is_secondhand_default=(name == "中古機"),
            )
            cat_map[name] = cat
        self.stdout.write(f"  建立 {len(cat_map)} 個類別")
        return cat_map

    # ---- 匯入商品 / 庫存 ----
    def _import(self, tenant, rows, mapping, wh_map, cat_map):
        used_names = set()

        def uniq_name(base, code):
            name = base or code
            if name not in used_names:
                used_names.add(name)
                return name
            name2 = f"{name} ({code})"
            i = 2
            while name2 in used_names:
                name2 = f"{name} ({code}-{i})"
                i += 1
            used_names.add(name2)
            return name2

        n_serial = n_bal = 0
        acc_products = {}  # 品號 -> Product

        # 中古機:逐列建商品 + 序號
        for r in rows:
            if r["prefix"] != SECONDHAND_PREFIX:
                continue
            cat = cat_map[mapping[r["prefix"]]]
            display = r["name"].split("/", 1)[1] if "/" in r["name"] else r["name"]
            p = Product(
                tenant=tenant,
                category=cat,
                name=uniq_name(display, r["code"]),
                barcode="",
                requires_serial=True,
                is_secondhand=True,
                is_active=True,
            )
            p.save()
            ProductSerial.objects.create(
                tenant=tenant,
                product=p,
                warehouse=wh_map[r["store"]],
                serial_no=r["code"],
                status=ProductSerial.Status.IN_STOCK,
                condition_note=r["name"],
            )
            n_serial += 1

        # 配件:依品號去重建商品,再各倉鋪 balance
        for r in rows:
            if r["prefix"] == SECONDHAND_PREFIX:
                continue
            code = r["code"]
            if code not in acc_products:
                cat = cat_map[mapping[r["prefix"]]]
                display = (
                    r["name"].split("/", 1)[1] if "/" in r["name"] else r["name"]
                )
                p = Product(
                    tenant=tenant,
                    category=cat,
                    name=uniq_name(display, code),
                    barcode=code,
                    requires_serial=False,
                    is_secondhand=False,
                    is_active=True,
                )
                p.save()
                acc_products[code] = p
            p = acc_products[code]
            if r["qty"] > 0:
                bal, created = StockBalance.objects.get_or_create(
                    tenant=tenant,
                    product=p,
                    warehouse=wh_map[r["store"]],
                    defaults={"qty": 0},
                )
                bal.qty += r["qty"]
                bal.save(update_fields=["qty"])
                n_bal += 1

        return len(acc_products) + n_serial, n_serial, n_bal
