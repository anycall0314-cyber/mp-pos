"""商品 CSV / Excel 匯入服務。

欄位對應(headers 可中文或英文 / 大小寫):
- 必填:品名(name)、類別(category,可填名稱或代碼)、品號(sku)
- 選填:安全庫存(safety_stock 預設 0)、建議售價(list_price)、條碼(barcode)

匯入規則:
- 類別不存在時自動建立(不用報錯)
- 已存在的品號 / 品名跳過
- 必填欄位空 → 跳過
- 所有新商品 lifecycle_status='pending'(待補齊),不影響庫存警示推論
- accessory_type 沿用 model default 'none';相容機型 / 屬性等預設空白
"""
from __future__ import annotations

import csv
import io
from io import BytesIO
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from .models import Category, Product

# 標題列關鍵字 → 內部欄位
_FIELD_MAP: dict[str, str] = {
    # name
    "品名": "name",
    "商品名稱": "name",
    "名稱": "name",
    "name": "name",
    # category
    "類別": "category",
    "類別名稱": "category",
    "類別代碼": "category",
    "category": "category",
    # sku
    "品號": "sku",
    "sku": "sku",
    # safety_stock
    "安全庫存": "safety_stock",
    "safety_stock": "safety_stock",
    "safety": "safety_stock",
    # list_price
    "建議售價": "list_price",
    "建議零售價": "list_price",
    "售價": "list_price",
    "list_price": "list_price",
    # barcode
    "條碼": "barcode",
    "barcode": "barcode",
}


def _norm_header(h: Any) -> str:
    return str(h or "").strip().lower()


def _normalize_row(raw: dict) -> dict:
    """標題對應到內部欄位 key,忽略不認識的 column。"""
    out: dict = {}
    for header, val in raw.items():
        target = _FIELD_MAP.get(_norm_header(header)) or _FIELD_MAP.get(str(header or "").strip())
        if target:
            out[target] = val
    return out


def _read_xlsx(file_obj) -> list[dict]:
    wb = load_workbook(
        filename=BytesIO(file_obj.read()), read_only=True, data_only=True
    )
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    header = [str(h or "").strip() for h in header]
    out = []
    for row in rows_iter:
        if row is None:
            continue
        d = {}
        for i, val in enumerate(row):
            if i < len(header) and header[i]:
                d[header[i]] = val
        # 整列空白跳過
        if not any((v is not None and str(v).strip()) for v in d.values()):
            continue
        out.append(d)
    return out


def _read_csv(file_obj) -> list[dict]:
    content = file_obj.read()
    if isinstance(content, bytes):
        # 去 BOM + decode
        if content.startswith(b"\xef\xbb\xbf"):
            content = content[3:]
        text = content.decode("utf-8-sig", errors="replace")
    else:
        text = content
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def _next_category_code(existing_codes: set[str]) -> str:
    """匯入時為新類別自動產生代碼:從 001 流水,避開現有代碼。"""
    n = 1
    while True:
        code = f"{n:03d}"
        if code not in existing_codes:
            return code
        n += 1


def _to_int(val, default=0) -> int:
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _to_decimal_str(val) -> str:
    if val is None or val == "":
        return "0"
    try:
        return str(float(val))
    except (TypeError, ValueError):
        return "0"


@transaction.atomic
def import_products_from_file(tenant, file_obj, filename: str, dry_run: bool = True) -> dict:
    """主入口:解析檔案 + 匯入。dry_run=True 不寫入資料庫(用於預覽)。"""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in ("xlsx", "xls"):
        raw_rows = _read_xlsx(file_obj)
    elif ext == "csv":
        raw_rows = _read_csv(file_obj)
    else:
        raise ValueError(f"不支援的檔案格式 .{ext}(請使用 .xlsx 或 .csv)")

    # 預載既有 cache
    cats = list(Category.objects.for_tenant(tenant))
    cat_by_code = {c.code.strip().lower(): c for c in cats}
    cat_by_name = {c.name.strip(): c for c in cats}
    existing_codes = {c.code for c in cats}
    existing_skus = set(
        Product.objects.for_tenant(tenant).values_list("sku", flat=True)
    )
    existing_names = set(
        Product.objects.for_tenant(tenant).values_list("name", flat=True)
    )

    success_rows: list[dict] = []
    skip_rows: list[dict] = []
    created_categories: list[str] = []

    sid = transaction.savepoint()

    for idx, raw in enumerate(raw_rows, start=2):  # 第 1 行是 header
        norm = _normalize_row(raw)
        name = (str(norm.get("name") or "")).strip()
        sku = (str(norm.get("sku") or "")).strip()
        category_raw = (str(norm.get("category") or "")).strip()

        if not name or not sku or not category_raw:
            skip_rows.append(
                {
                    "row_no": idx,
                    "sku": sku,
                    "name": name,
                    "reason": "必填欄位不齊(品名 / 類別 / 品號 至少缺一)",
                }
            )
            continue

        if sku in existing_skus:
            skip_rows.append(
                {
                    "row_no": idx,
                    "sku": sku,
                    "name": name,
                    "reason": f"品號 {sku} 已存在",
                }
            )
            continue
        if name in existing_names:
            skip_rows.append(
                {
                    "row_no": idx,
                    "sku": sku,
                    "name": name,
                    "reason": f"品名「{name}」已存在",
                }
            )
            continue

        # 類別:先試代碼(精確)、再試名稱;都找不到 → 建立
        cat = cat_by_code.get(category_raw.lower()) or cat_by_name.get(category_raw)
        if cat is None:
            code = _next_category_code(existing_codes)
            cat = Category.objects.create(
                tenant=tenant, code=code, name=category_raw
            )
            cat_by_code[code.lower()] = cat
            cat_by_name[category_raw] = cat
            existing_codes.add(code)
            created_categories.append(category_raw)

        safety = max(_to_int(norm.get("safety_stock"), 0), 0)
        list_price = _to_decimal_str(norm.get("list_price"))
        barcode = (str(norm.get("barcode") or "")).strip()

        p = Product(
            tenant=tenant,
            sku=sku,
            name=name,
            spec="",
            category=cat,
            list_price=list_price,
            safety_stock=safety,
            barcode=barcode,
            lifecycle_status=Product.LifecycleStatus.PENDING,
            is_active=True,
        )
        p.save()

        existing_skus.add(sku)
        existing_names.add(name)
        success_rows.append(
            {
                "row_no": idx,
                "sku": sku,
                "name": name,
                "category": category_raw,
            }
        )

    if dry_run:
        transaction.savepoint_rollback(sid)
    else:
        transaction.savepoint_commit(sid)

    return {
        "dry_run": dry_run,
        "total_rows": len(raw_rows),
        "success_count": len(success_rows),
        "skip_count": len(skip_rows),
        "created_categories": list(dict.fromkeys(created_categories)),
        "success_rows": success_rows[:200],
        "skip_rows": skip_rows[:200],
    }
