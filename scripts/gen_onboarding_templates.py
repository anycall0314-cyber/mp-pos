"""產生新經銷商上線用的 5 個 Excel 範本。

用法:
    cd /path/to/MP_POS系統
    backend/.venv/bin/python scripts/gen_onboarding_templates.py

輸出 5 個 xlsx 檔到 docs/templates/。

格式約定:
- 標題列在第 1 列,粗體 + 淺灰底色
- 必填欄位在標題後加「*」
- 第 2~4 列為範例資料,文字偏灰提示「示範用,匯入時請刪除」
- 凍結首列 + 適當欄寬
- 每個檔案另開一個工作表「填寫說明」放欄位解釋

若日後欄位需要調整,改本檔再重跑即可。
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "docs" / "templates"


HEADER_FILL = PatternFill(
    start_color="E8EEF7", end_color="E8EEF7", fill_type="solid"
)
HEADER_FONT = Font(bold=True, size=12)
HINT_FONT = Font(color="888888", italic=True, size=11)
TITLE_FONT = Font(bold=True, size=14)


def _write_template(filename, sheet_name, columns, examples, instructions):
    """columns: list of (display_header, width, comment_or_none, is_required)
    examples: list of list[str|int]
    instructions: list of (field, explanation) for the 填寫說明 sheet
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 1. 標題列
    for idx, (header, width, comment, required) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = f"{header}*" if required else header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
        if comment:
            cell.comment = Comment(comment, "MP POS")

    # 2. 範例列
    for ex_row_idx, ex in enumerate(examples, start=2):
        for col_idx, val in enumerate(ex, start=1):
            cell = ws.cell(row=ex_row_idx, column=col_idx)
            cell.value = val
            cell.font = HINT_FONT

    # 第 N+1 列加備註提示
    note_row = len(examples) + 2
    ws.cell(row=note_row, column=1).value = "↑ 上面 2-3 列為範例,請整列刪除後填入您的資料"
    ws.cell(row=note_row, column=1).font = HINT_FONT
    ws.merge_cells(
        start_row=note_row, start_column=1, end_row=note_row, end_column=len(columns)
    )

    # 3. 凍結首列
    ws.freeze_panes = "A2"

    # 4. 填寫說明分頁
    inst_ws = wb.create_sheet("填寫說明")
    inst_ws.cell(row=1, column=1).value = f"{sheet_name} — 欄位說明"
    inst_ws.cell(row=1, column=1).font = TITLE_FONT
    inst_ws.column_dimensions["A"].width = 22
    inst_ws.column_dimensions["B"].width = 80
    inst_ws.cell(row=3, column=1).value = "欄位"
    inst_ws.cell(row=3, column=2).value = "說明"
    inst_ws.cell(row=3, column=1).font = HEADER_FONT
    inst_ws.cell(row=3, column=2).font = HEADER_FONT
    inst_ws.cell(row=3, column=1).fill = HEADER_FILL
    inst_ws.cell(row=3, column=2).fill = HEADER_FILL
    for i, (field, explanation) in enumerate(instructions, start=4):
        inst_ws.cell(row=i, column=1).value = field
        inst_ws.cell(row=i, column=2).value = explanation
        inst_ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / filename
    wb.save(out_path)
    print(f"✓ {out_path}")


def build_01_products():
    columns = [
        ("品號", 16, "您舊系統的品項唯一代碼,例 PH-000123", True),
        ("品名", 36, "完整品名,例 iPhone 15 Pro 128G 黑色", True),
        ("規格", 20, "若品名已包含規格可留空", False),
        ("類別", 14, "大分類,例 手機 / 配件 / 中古機", True),
        ("條碼", 18, "原廠條碼 EAN-13,有則填", False),
        ("建議零售價", 14, "含稅單價,沒有可留空", False),
    ]
    examples = [
        ["PH-000123", "iPhone 15 Pro 128G 黑色", "128G / 黑色", "手機", "4710123456789", 36900],
        ["CH-000035", "Apple 20W USB-C 充電器", "20W", "配件", "194252084007", 590],
        ["AA-000007", "iPhone 13 Pro 256G 綠 (中古)", "256G / 綠色 / A 級", "中古機", "", 19900],
    ]
    instructions = [
        ("品號", "您舊 POS 系統的品項唯一代碼。新系統會直接沿用作為 SKU,因此**不可重複**。"),
        ("品名", "完整品名。建議格式『品牌 機型 容量 顏色』,有助於後續搜尋。"),
        ("規格", "若品名已包含規格(如容量/顏色)可留空。否則填入主要規格描述。"),
        ("類別", "大分類即可,例:手機、平板、配件、中古機、虛擬商品。新系統會自動建立未存在的類別。"),
        ("條碼", "原廠 EAN-13 條碼。配件商品有則必填,可用於掃描銷貨。"),
        ("建議零售價", "含稅單價(整數)。新系統會自動依稅率計算未稅。空白視為 0。"),
    ]
    _write_template(
        "01_商品主檔.xlsx", "商品主檔", columns, examples, instructions
    )


def build_02_serials():
    columns = [
        ("品號", 16, "對應商品主檔的品號", True),
        ("IMEI / 序號", 22, "該機唯一序號", True),
        ("倉別", 14, "該機現在所在門市,例 南大店", True),
        ("中古機成色", 12, "S/A/B/C/D,僅中古機需填", False),
        ("中古機售價", 14, "該支自訂售價,僅中古機需填", False),
        ("電池健康度", 12, "百分比數字,僅中古機需填", False),
        ("進貨成本", 14, "該機原始進貨成本,可選填", False),
    ]
    examples = [
        ["PH-000123", "354123456789012", "南大店", "", "", "", 28000],
        ["PH-000123", "354123456789999", "北倉", "", "", "", 28000],
        ["AA-000007", "AA-000007-A1", "南大店", "A", 19900, 95, 14500],
    ]
    instructions = [
        ("品號", "對應商品主檔中的品號。系統會用此鍵把序號掛到對應商品。"),
        ("IMEI / 序號", "手機 IMEI 15 碼,或廠牌自訂序號。中古機若無 IMEI 可用『品號-流水』格式。**整份檔案內不可重複**。"),
        ("倉別", "該機現在實際所在的門市名稱。系統會把同名稱的視為同一倉。"),
        ("中古機成色", "僅中古機需填:S(極新)/A(良好)/B(尚可)/C(普通)/D(瑕疵)。"),
        ("中古機售價", "該支獨立定價,銷貨時自動帶入。新機可留空(用商品主檔的建議零售價)。"),
        ("電池健康度", "整數百分比,例 95 代表 95%。僅中古機填寫。"),
        ("進貨成本", "該機進貨時的成本(未稅整數)。會用於毛利計算。空白系統會用 0,後續會吃虧計算毛利。"),
    ]
    _write_template(
        "02_庫存_序號.xlsx", "序號庫存", columns, examples, instructions
    )


def build_03_balances():
    columns = [
        ("品號", 16, "對應商品主檔", True),
        ("倉別", 14, "庫存所在門市", True),
        ("數量", 10, "在庫數(整數)", True),
    ]
    examples = [
        ["CH-000035", "南大店", 12],
        ["CH-000035", "北倉", 5],
        ["GL-000088", "南大店", 30],
    ]
    instructions = [
        ("品號", "對應商品主檔。配件 / 玻璃貼 / 充電線 等不追蹤序號的商品填這份。"),
        ("倉別", "庫存所在門市名稱。同品號可有多個倉,各佔一行。"),
        ("數量", "目前實際在庫數,整數。"),
    ]
    _write_template(
        "03_庫存_配件.xlsx", "配件庫存", columns, examples, instructions
    )


def build_04_members():
    columns = [
        ("姓名", 14, "會員姓名", True),
        ("電話", 14, "10 碼手機號,系統用此做唯一鍵", True),
        ("身分證", 14, "開門號 / 大額交易需要", False),
        ("生日", 14, "YYYY-MM-DD 格式", False),
        ("地址", 36, "完整地址", False),
        ("備註", 20, "例:VIP / 中古機買家", False),
    ]
    examples = [
        ["王小明", "0925123456", "A123456789", "1988-03-22", "新竹市東區光復路 100 號", "VIP"],
        ["林美玲", "0911222333", "", "", "", ""],
        ["陳大華", "0933888999", "B223456789", "1975-12-01", "新竹市北區中華路 5 段 88 號", "中古機買家"],
    ]
    instructions = [
        ("姓名", "會員姓名。"),
        ("電話", "10 碼手機號(無空白無分隔)。**同電話視為同會員**,系統以此去重。"),
        ("身分證", "開門號 / 中古機大額交易會需要。沒有可留空。"),
        ("生日", "YYYY-MM-DD 格式,例 1988-03-22。可留空。"),
        ("地址", "完整地址,可留空。"),
        ("備註", "自由文字,標記特殊客戶。"),
    ]
    _write_template(
        "04_會員主檔.xlsx", "會員主檔", columns, examples, instructions
    )


def build_05_purchases():
    columns = [
        ("會員電話", 14, "對應會員主檔", True),
        ("品號", 16, "對應商品主檔", True),
        ("數量", 8, "整數", True),
        ("單價", 12, "含稅單價(整數)", True),
        ("銷貨日期", 14, "YYYY-MM-DD", True),
        ("原單號", 16, "舊系統的銷貨單號,追溯用", False),
        ("IMEI / 序號", 22, "若是有序號商品,該支 IMEI", False),
    ]
    examples = [
        ["0925123456", "PH-000123", 1, 36900, "2024-08-12", "A-00012345", "354123456789012"],
        ["0925123456", "CH-000035", 1, 590, "2024-08-12", "A-00012345", ""],
        ["0911222333", "GL-000088", 2, 250, "2024-09-03", "A-00012450", ""],
    ]
    instructions = [
        ("會員電話", "對應會員主檔的電話。系統會以此找到對應會員。電話不存在於會員主檔時該筆會被略過。"),
        ("品號", "對應商品主檔的品號。品號不存在時該筆會被略過。"),
        ("數量", "正整數。退貨請以負數表示(若舊系統能匯出)。"),
        ("單價", "含稅單價(整數)。系統以此值顯示『前次成交價』供新銷貨時自動帶價。"),
        ("銷貨日期", "YYYY-MM-DD 格式。"),
        ("原單號", "舊系統的銷貨單號,僅作追溯用,不會還原成新系統的銷貨單。"),
        ("IMEI / 序號", "該筆若為手機 / 中古機,填當時賣出的 IMEI 號。可留空。"),
    ]
    _write_template(
        "05_消費紀錄.xlsx", "消費紀錄", columns, examples, instructions
    )


def main():
    build_01_products()
    build_02_serials()
    build_03_balances()
    build_04_members()
    build_05_purchases()
    print(f"\n完成。範本存在:{OUTPUT_DIR}")


if __name__ == "__main__":
    main()
